import openpyxl
path = "C:\\Users\\PC\\Documents\\myxlproject.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
roll = ["Manager1","Admin","Staff1","Student"]
user = ["xmngr1234",'xadmin1234',"xstaff@4567","xstud@987"]
password = ["a34555",'b5656565','c787878',"d977675"]
for r in range(1,len(roll)+1):
    sheet.cell(row=rows+r,column=1).value=(rows-1)+r
    sheet.cell(row=rows+r,column=2).value=roll[r-1]
    sheet.cell(row=rows+r,column=3).value=user[r-1]
    sheet.cell(row=rows+r,column=4).value=password[r-1]
workbook.save(path)

# import openpyxl
# path = "C:\\Users\\PC\\Documents\\myxlproject.xlsx"
# workbook = openpyxl.load_workbook(path)
# sheet = workbook.active
# rows = sheet.max_row
# cols = sheet.max_column
# print(rows)
# print(cols)
# for i in range(1,rows):
#     print(sheet.cell(row=i+1,column=1).value,end="_")
#     print(sheet.cell(row=i + 1, column=1).value,end="_")
#     print(sheet.cell(row=i + 1, column=1).value,end="_")
#     print(sheet.cell(row=i + 1, column=1).value, end="_")
#     print()
